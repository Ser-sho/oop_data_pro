from __future__ import annotations
from pathlib import Path
from typing import Any
import pandas as pd
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Font, PatternFill, Alignment


def _safe_sheet(name: str) -> str:
    for ch in '[]:*?/\\': name = name.replace(ch, '_')
    return name[:31]


def _write_df(writer, df: pd.DataFrame, sheet: str):
    if df is None or df.empty:
        pd.DataFrame({"Note":["No records available for this section."]}).to_excel(writer, sheet_name=_safe_sheet(sheet), index=False)
    else:
        df.to_excel(writer, sheet_name=_safe_sheet(sheet), index=False)


def _style_book(wb):
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for ws in wb.worksheets:
        ws.freeze_panes="A2"
        ws.auto_filter.ref=ws.dimensions
        for cell in ws[1]:
            cell.font=Font(bold=True)
            cell.fill=header_fill
            cell.alignment=Alignment(vertical="center")
        for col in ws.columns:
            max_len=max((len(str(c.value)) if c.value is not None else 0) for c in list(col)[:200])
            ws.column_dimensions[col[0].column_letter].width=min(max(max_len+2,10),45)


def _add_chart_sheet(writer, analysis):
    wb=writer.book
    ws=wb.create_sheet("Charts")
    ws["A1"]="Analytical charts"
    ws["A1"].font=Font(bold=True,size=16)
    row=3
    specs=[
        ("Channel mix", analysis.get("channel"), "Channel", "cases", "Channel", "Cases", "BarChart"),
        ("Top service demand", analysis.get("category1"), "value", "cases", "Case Category 1", "Cases", "BarChart"),
        ("Hourly case activity", analysis.get("dates",{}).get("by_hour"), "hour", "cases", "Hour of Day", "Cases", "BarChart"),
        ("Cases by reporting date", analysis.get("dates",{}).get("by_date"), "date", "cases", "Reporting Date", "Cases", "LineChart"),
    ]
    for title, df, xcol, ycol, xlabel, ylabel, typ in specs:
        if df is None or not isinstance(df,pd.DataFrame) or df.empty or xcol not in df.columns or ycol not in df.columns:
            continue
        data=df.head(12).copy()
        start=row
        ws.cell(row=row,column=1,value=title).font=Font(bold=True,size=13); row+=1
        ws.cell(row=row,column=1,value=xlabel); ws.cell(row=row,column=2,value=ylabel)
        for i,(_,r) in enumerate(data.iterrows(),start=row+1):
            ws.cell(i,1,value=r[xcol]); ws.cell(i,2,value=float(r[ycol]))
        if typ=="LineChart": chart=LineChart()
        else: chart=BarChart()
        chart.title=title
        chart.y_axis.title=ylabel
        chart.x_axis.title=xlabel
        chart.height=7; chart.width=13
        chart.add_data(Reference(ws,min_col=2,min_row=row,max_row=row+len(data)),titles_from_data=True)
        chart.set_categories(Reference(ws,min_col=1,min_row=row+1,max_row=row+len(data)))
        chart.legend=None
        if typ=="BarChart": chart.type="col"
        chart.dataLabels=DataLabelList(); chart.dataLabels.showVal=True
        ws.add_chart(chart, f"D{start}")
        row += len(data)+18
    ws.column_dimensions["A"].width=28; ws.column_dimensions["B"].width=16


def generate_addendum(output_path: str|Path, source_filename: str, sheet_name: str, analysis: dict[str,Any], intelligence: dict[str,Any], municipality: str, reporting_date, period_covered: str, close_time: str, voc_analysis=None) -> Path:
    output_path=Path(output_path); output_path.parent.mkdir(parents=True,exist_ok=True)
    s=analysis["summary"]
    context=pd.DataFrame([
        ["Report","OOP Corridor Daily Operations Report"],["Municipality / corridor",municipality],["Reporting date",reporting_date.isoformat()],
        ["Period covered",period_covered or "Not specified"],["Data close time",close_time],["Source file",source_filename],["Source sheet",sheet_name],
        ["Records analysed",s["records"]],["Valid cases",s["valid_cases"]],["Configured total wards",s["total_wards"] if s["total_wards"] is not None else "Not configured"],
        ["Distinct wards represented",s["distinct_wards"]],["Represented ward coverage %",s["ward_coverage_pct"] if s["ward_coverage_pct"] is not None else "Not configured"],
        ["Missing/unrepresented wards",s["missing_wards"] if s["missing_wards"] is not None else "Not configured"],
        ["Ward master listed allocations",s.get("ward_master_listed_allocations", "Not supplied")],
        ["Ward master unique ward numbers",s.get("ward_master_unique_wards", "Not supplied")],
        ["Ward master duplicate allocation rows",s.get("ward_master_duplicate_wards", "Not supplied")],
        ["Observed unique wards outside master",s.get("ward_master_outside_unique_wards", "Not supplied")],
        ["Daily cases",analysis.get("dates",{}).get("selected_day",{}).get("cases", "Not supplied")],
        ["Daily wards",analysis.get("dates",{}).get("selected_day",{}).get("daily_wards", "Not supplied")],
        ["New wards today",analysis.get("dates",{}).get("selected_day",{}).get("new_wards", "Not supplied")],
        ["Running wards",analysis.get("dates",{}).get("selected_day",{}).get("running_wards", "Not supplied")],
        ["Still needed",analysis.get("dates",{}).get("selected_day",{}).get("still_needed", "Not supplied")],
        ["Suggested new wards per remaining working day",analysis.get("dates",{}).get("selected_day",{}).get("suggested_new_wards_per_remaining_day", "Not supplied")],
        ["Remaining working days in week",analysis.get("dates",{}).get("selected_day",{}).get("remaining_workdays_in_week", "Not supplied")],
        ["Ward history status",analysis.get("dates",{}).get("selected_day",{}).get("ward_history_status", "Not supplied")],
        ["Retained ward history days",s.get("retained_ward_history_days", 0)],
        ["VOC responses",voc_analysis.get("responses") if voc_analysis else "Not supplied"],
    ],columns=["Item","Value"])
    dictionary=pd.DataFrame([{"field_role":k,"source_column":v,"usable":bool(v)} for k,v in analysis["columns"].items()])
    calculations=pd.DataFrame([
        {"metric":"Records analysed","value":s["records"],"calculation":"Number of rows in selected sheet"},
        {"metric":"Valid cases","value":s["valid_cases"],"calculation":"Rows with a non-empty mapped Case Number"},
        {"metric":"Invalid case records","value":s["invalid_case_records"],"calculation":"Records analysed - valid cases"},
        {"metric":"Distinct wards represented","value":s["distinct_wards"],"calculation":"Unique non-empty Ward Id values"},
        {"metric":"Ward coverage %","value":s["ward_coverage_pct"],"calculation":"Distinct wards / configured total wards × 100"},
        {"metric":"Missing/unrepresented wards","value":s["missing_wards"],"calculation":"Configured allocation benchmark - cumulative distinct wards reached through selected reporting date"},
        {"metric":"Daily cases","value":analysis.get("dates",{}).get("selected_day",{}).get("cases"),"calculation":"Valid case rows created on the manually selected reporting date"},
        {"metric":"Daily wards","value":analysis.get("dates",{}).get("selected_day",{}).get("daily_wards"),"calculation":"Distinct normalized wards with valid cases on the selected reporting date"},
        {"metric":"New wards","value":analysis.get("dates",{}).get("selected_day",{}).get("new_wards"),"calculation":"Selected-day wards not previously observed in retained reporting-week ward history"},
        {"metric":"Running wards","value":analysis.get("dates",{}).get("selected_day",{}).get("running_wards"),"calculation":"Distinct normalized wards observed cumulatively from the Monday of the reporting week through the selected reporting date"},
        {"metric":"Still needed","value":analysis.get("dates",{}).get("selected_day",{}).get("still_needed"),"calculation":"Configured ward allocation benchmark - cumulative running wards"},
        {"metric":"Suggested new wards per remaining working day","value":analysis.get("dates",{}).get("selected_day",{}).get("suggested_new_wards_per_remaining_day"),"calculation":"Ceiling of still needed divided by remaining Monday-Friday working days in the reporting week"},
        {"metric":"Observed unique wards outside master","value":s.get("ward_master_outside_unique_wards"),"calculation":"Distinct normalized case wards that do not match the supplied ward master"},
    ])
    with pd.ExcelWriter(output_path,engine="openpyxl") as writer:
        context.to_excel(writer,sheet_name="Report Context",index=False); calculations.to_excel(writer,sheet_name="Calculations",index=False); dictionary.to_excel(writer,sheet_name="Data Dictionary",index=False)
        _write_df(writer,analysis.get("quality"),"Data Quality")
        for key,sheet in [("status","Status"),("channel","Channel"),("case_type","Case Type"),("priority","Priority"),("city","City Area"),("category1","Demand Cat 1"),("category2","Demand Cat 2"),("category3","Demand Cat 3"),("owner","Owner")]: _write_df(writer,analysis.get(key),sheet)
        dates=analysis.get("dates",{}); _write_df(writer,dates.get("by_date") if dates.get("available") else None,"Cases by Date"); _write_df(writer,dates.get("by_hour") if dates.get("available") else None,"Cases by Hour")
        _write_df(writer,dates.get("daily_tracker") if dates.get("available") else None,"Daily Tracker")
        _write_df(writer,analysis.get("ward_mapping"),"Ward Mapping QA")
        _write_df(writer,dates.get("ward_coverage") if dates.get("available") else None,"Ward Coverage Detail")
        _write_df(writer,dates.get("missing_wards") if dates.get("available") else None,"Missing Wards")
        _write_df(writer,analysis.get("corridor_coverage"),"Corridor Coverage")
        _write_df(writer,analysis.get("cca_coverage"),"CCA Coverage")
        _write_df(writer,analysis.get("cca_daily_tracker"),"CCA Daily Tracker")
        _write_df(writer,analysis.get("channel_new_wards"),"Channel New Wards")
        if analysis.get("ward_master",{}).get("available"):
            _write_df(writer,analysis["ward_master"].get("duplicate_wards"),"Ward Master Exceptions")
        _write_df(writer,intelligence.get("findings"),"Findings"); _write_df(writer,intelligence.get("recommendations"),"Recommendations"); _write_df(writer,intelligence.get("evidence"),"Evidence Register"); _write_df(writer,intelligence.get("evidence_details"),"Evidence Detail"); _write_df(writer,intelligence.get("qa"),"Traceability QA"); _write_df(writer,intelligence.get("final_report_qa"),"Final Report QA")
        sd=analysis.get("dates",{}).get("selected_day",{})
        ward_note=pd.DataFrame({"Item":["Ward representation through selected reporting week"],"Configured allocation benchmark":[s["total_wards"]],"Cumulative wards reached":[sd.get("running_wards",s["distinct_wards"])],"Coverage %":[sd.get("coverage_pct",s["ward_coverage_pct"])],"Still needed":[sd.get("still_needed",s["missing_wards"])],"Master unique ward numbers":[s.get("ward_master_unique_wards")]}); ward_note.to_excel(writer,sheet_name="Ward Coverage",index=False)
        if voc_analysis: _write_df(writer,pd.DataFrame([voc_analysis]),"VOC Summary")
        _add_chart_sheet(writer,analysis)
        _style_book(writer.book)
    return output_path
